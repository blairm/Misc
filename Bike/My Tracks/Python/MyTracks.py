import sys

import csv

import math
import numpy

import datetime

import utm
from geopy import distance

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
)


class MyTracksPoint:
    def __init__( self,
                    latitude:float = 0,
                    longitude:float = 0,
                    east:float = 0,
                    north:float = 0,
                    altitude:float = 0,
                    bearing:float = 0,
                    accuracy:float = 0,
                    speed:float = 0,
                    time:float = 0 ):
        self.latitude:float = latitude
        self.longitude:float = longitude
        self.east:float = east
        self.north:float = north
        self.altitude:float = altitude
        self.bearing:float = bearing
        self.accuracy:float = accuracy
        self.speed:float = speed
        self.time:float = time


class InfoPoint:
    def __init__( self,
                    east:float = 0,
                    north:float = 0,
                    altitude:float = 0,
                    distance:float = 0,
                    accuracy:float = 0,
                    time:float = 0 ):
        self.east:float = east
        self.north:float = north
        self.altitude:float = altitude
        self.distance:float = distance
        self.accuracy:float = accuracy
        self.time:float = time


def read_float_or_zero( value:str ) -> float:
    result:float = 0
    if value != "":
        result = float( value )
    return result

def speed_ms_to_kph( speedMs:float ) -> float:
    result:float = speedMs * 3.6
    return result

def distance_m_to_km( distanceM:float ) -> float:
    result:float = distanceM * 0.001
    return result

def distance_km_to_m( distanceKM:float ) -> float:
    result:float = distanceKM * 1000
    return result


def set_numeric_cell( sheet:Worksheet, row:int, column:int, value, format:str ) -> Cell:
    result:Cell = sheet.cell( row, column, value )
    result.number_format = format
    return result

def set_cell_row( sheet:Worksheet, row:int, column:int, value:[] ):
    for i in range( 0, len( value ) ):
        sheet.cell( row, i + column, value[ i ] )


def add_marker( chart:ScatterChart, sheet:Worksheet, xRow:int, xColumn:int, yRow:int, yColumn:int, labelReference:str, colour:str ) -> Series:
    valuesX:Reference = Reference( sheet, min_col = xColumn, min_row = xRow, max_col = xColumn, max_row = xRow )
    valuesY:Reference = Reference( sheet, min_col = yColumn, min_row = yRow, max_col = yColumn, max_row = yRow )
    result:Series = Series( valuesY, valuesX )
    result.title = openpyxl.chart.series.SeriesLabel( openpyxl.chart.data_source.StrRef( labelReference ) )
    result.smooth = False;
    result.marker.symbol = "circle"
    result.marker.size = 8
    result.graphicalProperties.line.solidFill = colour
    chart.series.append( result )
    return result


myTracksPoint:MyTracksPoint = []

eastBounds:float = [ sys.float_info.max, sys.float_info.min ]
northBounds:float = [ sys.float_info.max, sys.float_info.min ]


filename:str = "../_test"

if len( sys.argv ) > 1 and sys.argv[ 1 ] != "":
    filename = sys.argv[ 1 ].rsplit( ".", 1 )[ 0 ]


with open( f"{filename}.csv" ) as file:
    reader:csv.reader = csv.reader( file, delimiter = ',' )
    
    segment:int = 1
    segmentTimeGap:float = 0
    startTime:datetime
    lastRowTime:datetime

    for row in reader:
        if len( row ) > 0 and row[ 0 ].isnumeric():
            latitude:float = read_float_or_zero( row[ 2 ] )
            longitude:float = read_float_or_zero( row[ 3 ] )


            east, north, zoneNumber, zoneLetter = utm.from_latlon( latitude, longitude )

            eastBounds[ 0 ] = min( eastBounds[ 0 ], east )
            eastBounds[ 1 ] = max( eastBounds[ 1 ], east )
            northBounds[ 0 ] = min( northBounds[ 0 ], north )
            northBounds[ 1 ] = max( northBounds[ 1 ], north )
            

            altitude:float = read_float_or_zero( row[ 4 ] )
            bearing:float = read_float_or_zero( row[ 5 ] )
            accuracy:float = read_float_or_zero( row[ 6 ] )
            speed:float = speed_ms_to_kph( read_float_or_zero( row[ 7 ] ) )


            rowTime:datetime = datetime.datetime.strptime( row[ 8 ], "%Y-%m-%dT%H:%M:%S.%fZ" )

            if len( myTracksPoint ) == 0:
                startTime = rowTime

            if int( row[ 0 ] ) > segment:
                segmentTimeGap += ( rowTime - lastRowTime ).total_seconds()
                segment = int( row[ 0 ] )

            pointTime:float = ( rowTime - startTime ).total_seconds() - segmentTimeGap

            
            myTracksPoint.append( MyTracksPoint( latitude, longitude, east, north, altitude, bearing, accuracy, speed, pointTime ) )

            lastRowTime = rowTime


myTracksPoint.sort( key=lambda MyTracksPoint: MyTracksPoint.time )


prevPoint:MyTracksPoint = myTracksPoint[ 0 ]
infoPoint:InfoPoint = [ InfoPoint( prevPoint.east,
                                    prevPoint.north,
                                    prevPoint.altitude,
                                    0,
                                    prevPoint.accuracy,
                                    prevPoint.time ) ]

for i in range( 1, len( myTracksPoint ) ):
    point:MyTracksPoint = myTracksPoint[ i ]

    testDist:float = distance.distance( ( point.latitude, point.longitude ), ( prevPoint.latitude, prevPoint.longitude ) ).m
    if testDist > point.accuracy + prevPoint.accuracy:
        infoPoint.append( InfoPoint( point.east,
                                        point.north,
                                        point.altitude,
                                        distance_m_to_km( testDist ) + infoPoint[ -1 ].distance,
                                        point.accuracy,
                                        point.time ) )
        prevPoint = point


workbook = Workbook()

dataSheet = workbook.active
dataSheet.title = "Data"

set_cell_row( dataSheet, 1, 1, [ "Latitude (degrees)",
                                    "Longitude (degrees)",
                                    "East (m)",
                                    "North (m)",
                                    "Altitude (m)",
                                    "Bearing (degrees)",
                                    "Accuracy (m)",
                                    "Speed (kph)",
                                    "Time" ] )


for i in range( 0, len( myTracksPoint ) ):
    dataSheet.cell( i + 2, 1, myTracksPoint[ i ].latitude )
    dataSheet.cell( i + 2, 2, myTracksPoint[ i ].longitude )
    set_numeric_cell( dataSheet, i + 2, 3, myTracksPoint[ i ].east, "0.00" )
    set_numeric_cell( dataSheet, i + 2, 4, myTracksPoint[ i ].north, "0.00" )
    set_numeric_cell( dataSheet, i + 2, 5, myTracksPoint[ i ].altitude, "0.00" )
    set_numeric_cell( dataSheet, i + 2, 6, myTracksPoint[ i ].bearing, "0.00" )
    set_numeric_cell( dataSheet, i + 2, 7, myTracksPoint[ i ].accuracy, "0.00" )
    set_numeric_cell( dataSheet, i + 2, 8, myTracksPoint[ i ].speed, "0.00" )
    set_numeric_cell( dataSheet, i + 2, 9, f"=TIME( 0, 0, {float(myTracksPoint[ i ].time)} )", "[h]:mm:ss" )


infoSheet = workbook.create_sheet( "Info" )

infoSheet.cell( 1, 1, "Total Distance (km)" )
set_numeric_cell( infoSheet, 1, 2,  infoPoint[ -1 ].distance, "0.00" )
infoSheet.cell( 2, 1, "Total Time" )
set_numeric_cell( infoSheet, 2, 2,  "=INDEX( $E$6:$E$65535, COUNT( $E$6:$E$65535 ) )", "[h]:mm:ss" )
infoSheet.cell( 3, 1, "Average Speed (kph)" )
set_numeric_cell( infoSheet, 3, 2,  speed_ms_to_kph( distance_km_to_m( infoPoint[ -1 ].distance ) / myTracksPoint[ -1 ].time ), "0.00" )


#start marker data
endIndex:int = 100
infoSheet.cell( 1, 4, "Start" )
infoSheet.cell( 1, 5, 0 )
infoSheet.cell( 2, 4, "Map" )
set_numeric_cell( infoSheet, 2, 5, f"=INDEX( $A$6:$A$65535, 1 + ROUND( ( COUNT( $A$6:$A$65535 ) - 1 ) * ( $E$1 / {endIndex} ) ) )", "0.00" )
set_numeric_cell( infoSheet, 2, 6, f"=INDEX( $B$6:$B$65535, 1 + ROUND( ( COUNT( $B$6:$B$65535 ) - 1 ) * ( $E$1 / {endIndex} ) ) )", "0.00" )
infoSheet.cell( 3, 4, "Altitude" )
set_numeric_cell( infoSheet, 3, 5, f"=INDEX( $D$6:$D$65535, 1 + ROUND( ( COUNT( $D$6:$D$65535 ) - 1 ) * ( $E$1 / {endIndex} ) ) )", "0.00" )
set_numeric_cell( infoSheet, 3, 6, f"=INDEX( $C$6:$C$65535, 1 + ROUND( ( COUNT( $C$6:$C$65535 ) - 1 ) * ( $E$1 / {endIndex} ) ) )", "0.00" )

#end marker data
infoSheet.cell( 1, 7, "End" )
infoSheet.cell( 1, 8, endIndex )
infoSheet.cell( 2, 7, "Map" )
set_numeric_cell( infoSheet, 2, 8, f"=INDEX( $A$6:$A$65535, 1 + ROUND( ( COUNT( $A$6:$A$65535 ) - 1 ) * ( $H$1 / {endIndex} ) ) )", "0.00" )
set_numeric_cell( infoSheet, 2, 9, f"=INDEX( $B$6:$B$65535, 1 + ROUND( ( COUNT( $B$6:$B$65535 ) - 1 ) * ( $H$1 / {endIndex} ) ) )", "0.00" )
infoSheet.cell( 3, 7, "Altitude" )
set_numeric_cell( infoSheet, 3, 8, f"=INDEX( $D$6:$D$65535, 1 + ROUND( ( COUNT( $D$6:$D$65535 ) - 1 ) * ( $H$1 / {endIndex} ) ) )", "0.00" )
set_numeric_cell( infoSheet, 3, 9, f"=INDEX( $C$6:$C$65535, 1 + ROUND( ( COUNT( $C$6:$C$65535 ) - 1 ) * ( $H$1 / {endIndex} ) ) )", "0.00" )


#selected map data
infoSheet.cell( 1, 11, "Distance (km)" )
set_numeric_cell( infoSheet, 2, 11, "=H3-E3", "0.00" )

infoSheet.cell( 1, 12, "Ascent (m)" )
set_numeric_cell( infoSheet, 2, 12, "=I3-F3", "0.00" )

infoSheet.cell( 1, 13, "Gradient" )
set_numeric_cell( infoSheet, 2, 13, "=L2 / ( K2 * 1000 )", "0.00%" )

infoSheet.cell( 1, 14, "Time" )
set_numeric_cell( infoSheet, 2, 14, f"=INDEX( $E$6:$E$65535, 1 + ROUND( ( COUNT( $E$6:$E$65535 ) - 1 ) * ( $H$1 / {endIndex} ) ) ) - INDEX( $E$6:$E$65535, 1 + ROUND( ( COUNT( $E$6:$E$65535 ) - 1 ) * ( $E$1 / {endIndex} ) ) )", "[h]:mm:ss" )

infoSheet.cell( 1, 15, "Speed (kph)" )
set_numeric_cell( infoSheet, 2, 15, "=K2 / ( HOUR( N2 ) + ( MINUTE( N2 ) / 60 ) + ( SECOND( N2 ) / 3600 ) )", "0.00" )

infoSheet.cell( 1, 16, "Rider Weight (kg)" )
set_numeric_cell( infoSheet, 2, 16, 65, "0.00" )

infoSheet.cell( 1, 17, "Kit Weight (kg)" )
set_numeric_cell( infoSheet, 2, 17, 5, "0.00" )

infoSheet.cell( 1, 18, "Bike Weight (kg)" )
set_numeric_cell( infoSheet, 2, 18, 9, "0.00" )

infoSheet.cell( 1, 19, "Total Weight (kg)" )
set_numeric_cell( infoSheet, 2, 19, "=SUM( P2:R2 )", "0.00" )

infoSheet.cell( 1, 20, "Rolling Resistance" )
infoSheet.cell( 2, 20, 0.009 )

infoSheet.cell( 1, 21, "CdA" )
infoSheet.cell( 2, 21, 0.408 )

infoSheet.cell( 1, 22, "Efficiency" )
set_numeric_cell( infoSheet, 2, 22, 0.95, "0.00%" )

infoSheet.cell( 1, 23, "Power (w)" )
#Fg = gravity * sin( atn( gradient ) ) * weight
#Frr = gravity * cos( atn( gradient ) ) * weight * rollingResistance
#airPressure = 1.225 * Exp( -0.00011856 * elevation )
#power = ( speed / efficiency ) / ( Fg + Frr + ( 0.5 * CdA * airPressure * speed^2 ) )
set_numeric_cell( infoSheet, 2, 23,
                    "==( ( O2 / 3.6 ) / V2 ) * ( ( 9.81 * SIN( ATAN( M2 ) ) * S2 ) + ( 9.81 * COS( ATAN( M2 ) ) * S2 * T2 ) + ( 0.5 * U2 * ( 1.225 * EXP( -0.00011856 * F3 ) ) * ( ( O2 / 3.6 )^2 ) ) )",
                    "0.00" )

infoSheet.cell( 1, 24, "W/Kg" )
set_numeric_cell( infoSheet, 2, 24, "=W2 / P2", "0.00" )


set_cell_row( infoSheet, 5, 1, [ "East (m)", "North (m)", "Altitude (m)", "Distance (km)", "Time" ] )
for i in range( 0, len( infoPoint ) ):
    set_numeric_cell( infoSheet, i + 6, 1, infoPoint[ i ].east, "0.00" )
    set_numeric_cell( infoSheet, i + 6, 2, infoPoint[ i ].north, "0.00" )
    set_numeric_cell( infoSheet, i + 6, 3, infoPoint[ i ].altitude, "0.00" )
    set_numeric_cell( infoSheet, i + 6, 4, infoPoint[ i ].distance, "0.00" )
    set_numeric_cell( infoSheet, i + 6, 5, f"=TIME( 0, 0, {float( infoPoint[ i ].time )} )", "[h]:mm:ss" )


mapChart = ScatterChart()
mapChart.legend = None
mapChart.x_axis.delete = True
mapChart.x_axis.majorGridlines = None
mapChart.x_axis.scaling.min = eastBounds[ 0 ]
mapChart.x_axis.scaling.max = eastBounds[ 1 ]
mapChart.y_axis.delete = True
mapChart.y_axis.majorGridlines = None
mapChart.y_axis.scaling.min = northBounds[ 0 ]
mapChart.y_axis.scaling.max = northBounds[ 1 ]
mapChart.width = mapChart.height * ( ( eastBounds[ 1 ] - eastBounds[ 0 ] ) / ( northBounds[ 1 ] - northBounds[ 0 ] ) )

mapValuesX = Reference( infoSheet, min_col = 1, min_row = 6, max_col = 1, max_row = len( infoPoint ) + 6 )
mapValuesY = Reference( infoSheet, min_col = 2, min_row = 6, max_col = 2, max_row = len( infoPoint ) + 6 )
mapSeries = Series( mapValuesY, mapValuesX )
mapSeries.title = openpyxl.chart.series.SeriesLabel( openpyxl.chart.data_source.StrRef( "'Info'!D2" ) )
mapSeries.smooth = False;
mapSeries.graphicalProperties.line.width = 0
mapSeries.graphicalProperties.line.solidFill = "2A6099"
mapChart.series.append( mapSeries )

add_marker( mapChart, infoSheet, 2, 5, 2, 6, "'Info'!D1", "00A933" )
add_marker( mapChart, infoSheet, 2, 8, 2, 9, "'Info'!G1", "BE4B48" )

infoSheet.add_chart( mapChart, "G6" )


altitudeChart = ScatterChart()
altitudeChart.legend = None
altitudeChart.x_axis.title = "Distance (Km)"
altitudeChart.x_axis.majorGridlines = None
altitudeChart.x_axis.scaling.min = 0
altitudeChart.x_axis.scaling.max = infoPoint[ -1 ].distance
altitudeChart.y_axis.title = "Altitude (m)"
altitudeChart.y_axis.majorGridlines = None

altitudeValuesX = Reference( infoSheet, min_col = 4, min_row = 6, max_col = 4, max_row = len( infoPoint ) + 6 )
altitudeValuesY = Reference( infoSheet, min_col = 3, min_row = 6, max_col = 3, max_row = len( infoPoint ) + 6 )
altitudeSeries = Series( altitudeValuesY, altitudeValuesX )
mapSeries.title = openpyxl.chart.series.SeriesLabel( openpyxl.chart.data_source.StrRef( "'Info'!D2" ) )
altitudeSeries.smooth = False;
altitudeSeries.graphicalProperties.line.width = 0
altitudeSeries.graphicalProperties.line.solidFill = "2A6099"
altitudeChart.series.append( altitudeSeries )

add_marker( altitudeChart, infoSheet, 3, 5, 3, 6, "'Info'!D1", "00A933" )
add_marker( altitudeChart, infoSheet, 3, 8, 3, 9, "'Info'!G1", "BE4B48" )

infoSheet.add_chart( altitudeChart, "G22" )


workbook.save( f"{filename}.xlsx" )